"""
Punting Club — Competition Format Financial Model
Compares 4 formats: Annual | Bi-Annual | Quarterly | Monthly
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

# ── Colour palette ────────────────────────────────────────────────────────────
DARK_BG   = "0F172A"
CARD      = "1E293B"
GREEN     = "22C55E"
GREEN_D   = "166534"
AMBER     = "F59E0B"
RED       = "EF4444"
BLUE      = "3B82F6"
PURPLE    = "8B5CF6"
SLATE     = "334155"
WHITE     = "F1F5F9"
MUTED     = "94A3B8"
YELLOW_HL = "FEF08A"  # highlight winner

MODEL_COLOURS = {
    "Annual":    "3B82F6",   # blue
    "Bi-Annual": "8B5CF6",   # purple
    "Quarterly": "F59E0B",   # amber
    "Monthly":   "22C55E",   # green  ← WINNER
}

YEARS       = [1, 2, 3, 4]
PUB_COUNTS  = [10, 30, 100, 300]
TEAMS_PER   = 10
AD_REV      = [2_000, 10_000, 50_000, 150_000]   # advertising per year

# ── Model definitions ─────────────────────────────────────────────────────────
MODELS = [
    dict(name="Annual",    comps=1,  weeks=52, buyin=1_000),
    dict(name="Bi-Annual", comps=2,  weeks=26, buyin=600),
    dict(name="Quarterly", comps=4,  weeks=13, buyin=350),
    dict(name="Monthly",   comps=12, weeks=4,  buyin=150),
]

# ── Cost building blocks ──────────────────────────────────────────────────────
# Base costs (same regardless of model)
BASE = dict(
    hosting   = [228,   1_188,  1_188,  6_000],
    database  = [300,     300,  7_188, 18_000],
    ai_api    = [500,   1_200,  4_000, 12_000],
    domain    = [ 50,      50,     50,    200],
    hr        = [  0,  20_000, 80_000,240_000],
    marketing = [2_000, 10_000, 30_000, 80_000],
    legal     = [3_000,  5_000, 10_000, 20_000],
    misc      = [  500,  2_000,  5_000, 15_000],
)

PAYMENT_RATE        = 0.0175   # 1.75 % of jackpot collected
DEPOSIT_RATE        = 0.04     # 4 % annual
COMP_OVERHEAD_RATES = [20, 15, 10, 8]   # $ per competition per pub by year

# ── Calculation helpers ───────────────────────────────────────────────────────
def calc_model(m, yi):
    """Return revenue and cost breakdown for a model at year-index yi."""
    pubs   = PUB_COUNTS[yi]
    comps  = m["comps"]
    weeks  = m["weeks"]
    buyin  = m["buyin"]
    teams  = TEAMS_PER

    # Revenue
    jackpot_per_comp = teams * buyin                    # per pub per comp
    annual_jackpot   = pubs * comps * jackpot_per_comp  # total gross collected
    commission       = annual_jackpot * 0.10
    deposit_per_comp = jackpot_per_comp * pubs * 0.90
    interest_per_comp = deposit_per_comp * DEPOSIT_RATE * (weeks / 52)
    interest_total   = interest_per_comp * comps
    advertising      = AD_REV[yi]
    total_revenue    = commission + interest_total + advertising

    # Base costs
    base_total = sum(v[yi] for v in BASE.values())

    # Model-specific costs
    payment_proc = annual_jackpot * PAYMENT_RATE
    comp_overhead = pubs * comps * COMP_OVERHEAD_RATES[yi]
    total_cost = base_total + payment_proc + comp_overhead

    net_profit = total_revenue - total_cost
    margin     = net_profit / total_revenue if total_revenue else 0

    return dict(
        pubs=pubs, comps=comps, buyin=buyin,
        jackpot_total=annual_jackpot,
        commission=commission,
        interest=interest_total,
        advertising=advertising,
        total_revenue=total_revenue,
        # cost breakdown
        infrastructure=sum(BASE[k][yi] for k in ["hosting","database","ai_api","domain"]),
        hr=BASE["hr"][yi],
        marketing=BASE["marketing"][yi],
        legal=BASE["legal"][yi],
        misc=BASE["misc"][yi],
        payment_proc=payment_proc,
        comp_overhead=comp_overhead,
        total_cost=total_cost,
        net_profit=net_profit,
        margin=margin,
    )


# Pre-compute all results
results = {}
for m in MODELS:
    results[m["name"]] = [calc_model(m, yi) for yi in range(4)]

# ── Excel helpers ─────────────────────────────────────────────────────────────
def fill(hex_colour):
    return PatternFill("solid", fgColor=hex_colour)

def bold(size=11, colour="000000"):
    return Font(bold=True, size=size, color=colour)

def reg(size=10, colour="000000"):
    return Font(size=size, color=colour)

def border_thin(sides="all"):
    s = Side(style="thin", color="CCCCCC")
    n = Side(style=None)
    if sides == "all":
        return Border(left=s, right=s, top=s, bottom=s)
    if sides == "bottom":
        return Border(bottom=s)
    return Border(left=s, right=s, top=s, bottom=s)

def money(ws, row, col, value, fmt="$#,##0", bg=None, bold_=False, colour="000000"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = border_thin()
    if bg:  cell.fill = fill(bg)
    cell.font = Font(bold=bold_, size=10, color=colour)
    return cell

def header_cell(ws, row, col, text, bg=DARK_BG, fg=WHITE, size=11, merge_to=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=True, size=size, color=fg)
    cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_thin()
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=merge_to)
    return cell

def label_cell(ws, row, col, text, bg="FFFFFF", fg="000000", indent=0):
    cell = ws.cell(row=row, column=col, value=("  " * indent) + text)
    cell.font = Font(size=10, color=fg)
    cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = border_thin()
    return cell

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ── WORKBOOK ──────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — COMPARISON SUMMARY
# ════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Comparison Summary"
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A4"

# Title
ws.row_dimensions[1].height = 36
ws.merge_cells("A1:R1")
t = ws["A1"]
t.value = "PUNTING CLUB — COMPETITION FORMAT FINANCIAL COMPARISON"
t.font = Font(bold=True, size=16, color=WHITE)
t.fill = fill(DARK_BG)
t.alignment = Alignment(horizontal="center", vertical="center")

ws.row_dimensions[2].height = 20
ws.merge_cells("A2:R2")
sub = ws["A2"]
sub.value = "Revenue · Cost · Net Profit analysis across 4 competition formats (Years 1–4)"
sub.font = Font(size=11, color=MUTED)
sub.fill = fill(DARK_BG)
sub.alignment = Alignment(horizontal="center", vertical="center")

# ── Section headers ───────────────────────────────────────────────────────
ws.row_dimensions[4].height = 28
header_cell(ws, 4, 1, "Metric", CARD, WHITE, 11)
col = 2
year_cols = {}   # (model_name, year) -> col index
model_col_start = {}
for m in MODELS:
    model_col_start[m["name"]] = col
    header_cell(ws, 4, col, m["name"], MODEL_COLOURS[m["name"]], WHITE, 11, col+3)
    for i, y in enumerate(YEARS):
        year_cols[(m["name"], y)] = col + i
    col += 4

set_col_width(ws, 1, 28)
for c in range(2, col):
    set_col_width(ws, c, 14)

ws.row_dimensions[5].height = 22
label_cell(ws, 5, 1, "Year", CARD, WHITE)
col = 2
for m in MODELS:
    for y in YEARS:
        cell = ws.cell(row=5, column=col, value=f"Year {y}")
        cell.font = Font(bold=True, size=10, color=WHITE)
        cell.fill = fill(MODEL_COLOURS[m["name"]])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin()
        col += 1

# ── Rows ──────────────────────────────────────────────────────────────────
SECTIONS = [
    ("INPUTS", None, [
        ("Pub Count",         "pubs",          False, "0"),
        ("Competitions / Yr", "comps",         False, "0"),
        ("Buy-in per Team",   "buyin",         False, "$#,##0"),
        ("Teams per Pub",     None,            False, "0"),   # constant
    ]),
    ("REVENUE", GREEN_D, [
        ("Gross Jackpot",     "jackpot_total",  False, "$#,##0"),
        ("Commission (10%)",  "commission",     False, "$#,##0"),
        ("Term Deposit Int.", "interest",       False, "$#,##0"),
        ("Advertising",       "advertising",    False, "$#,##0"),
        ("TOTAL REVENUE",     "total_revenue",  True,  "$#,##0"),
    ]),
    ("COSTS", "7F1D1D", [
        ("Infrastructure",   "infrastructure",  False, "$#,##0"),
        ("HR / Personnel",   "hr",              False, "$#,##0"),
        ("Marketing",        "marketing",       False, "$#,##0"),
        ("Legal/Compliance", "legal",           False, "$#,##0"),
        ("Misc",             "misc",            False, "$#,##0"),
        ("Payment Proc.",    "payment_proc",    False, "$#,##0"),
        ("Comp. Overhead",   "comp_overhead",   False, "$#,##0"),
        ("TOTAL COSTS",      "total_cost",      True,  "$#,##0"),
    ]),
    ("PROFIT", None, [
        ("NET PROFIT",        "net_profit",     True,  "$#,##0"),
        ("Margin %",          "margin",         True,  "0.0%"),
    ]),
]

row = 6
ALT1 = "F8FAFC"
ALT2 = "EEF2FF"

for sect_name, sect_colour, rows in SECTIONS:
    # Section title row
    ws.row_dimensions[row].height = 24
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col-1)
    sc = ws.cell(row=row, column=1, value=f"  {sect_name}")
    sc.font = Font(bold=True, size=11, color=WHITE)
    sc.fill = fill(sect_colour or SLATE)
    sc.alignment = Alignment(horizontal="left", vertical="center")
    sc.border = border_thin()
    row += 1

    for ri, (label, key, is_bold, fmt) in enumerate(rows):
        ws.row_dimensions[row].height = 20
        bg_row = ALT1 if ri % 2 == 0 else ALT2
        is_total = label.startswith("TOTAL") or label == "NET PROFIT" or label == "Margin %"
        row_bg = "DBEAFE" if (is_total and sect_name == "REVENUE") else \
                 "FEE2E2" if (is_total and sect_name == "COSTS") else \
                 "DCFCE7" if (is_total and sect_name == "PROFIT") else bg_row

        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(bold=is_bold, size=10, color="111827")
        lc.fill = fill(row_bg)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lc.border = border_thin()

        for m in MODELS:
            is_winner = m["name"] == "Monthly"
            for yi, y in enumerate(YEARS):
                c = year_cols[(m["name"], y)]
                if key is None:
                    val = TEAMS_PER
                else:
                    val = results[m["name"]][yi].get(key, 0)

                winner_bg = "D1FAE5" if (is_winner and is_total) else None
                bg = winner_bg or row_bg

                cell_fg = "065F46" if (is_winner and is_total and sect_name in ("PROFIT","REVENUE")) else \
                          "991B1B" if (is_total and sect_name == "COSTS" and not is_winner) else "111827"

                # Flag negative profit in red
                if key == "net_profit" and isinstance(val, (int, float)) and val < 0:
                    bg = "FEE2E2"; cell_fg = "991B1B"

                money(ws, row, c, val, fmt, bg, is_bold, cell_fg)
        row += 1

# ── Cumulative row ────────────────────────────────────────────────────────
ws.row_dimensions[row].height = 26
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col-1)
cum = ws.cell(row=row, column=1, value="  CUMULATIVE 4-YEAR NET PROFIT")
cum.font = Font(bold=True, size=12, color=WHITE)
cum.fill = fill(DARK_BG)
cum.alignment = Alignment(horizontal="left", vertical="center")
cum.border = border_thin()
row += 1

ws.row_dimensions[row].height = 24
lc = ws.cell(row=row, column=1, value="4-Year Total")
lc.font = Font(bold=True, size=11, color="111827")
lc.fill = fill("FEF9C3")
lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
lc.border = border_thin()

for m in MODELS:
    cum_val = sum(results[m["name"]][yi]["net_profit"] for yi in range(4))
    is_winner = m["name"] == "Monthly"
    bg = "D1FAE5" if is_winner else "FEF9C3"
    fg = "065F46" if is_winner else "111827"
    # Span all 4 year cols for this model
    start_c = year_cols[(m["name"], 1)]
    ws.merge_cells(start_row=row, start_column=start_c, end_row=row, end_column=start_c+3)
    cell = ws.cell(row=row, column=start_c, value=cum_val)
    cell.number_format = "$#,##0"
    cell.font = Font(bold=True, size=12, color=fg)
    cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_thin()
    if is_winner:
        ws.cell(row=row, column=start_c).value = f"★ ${cum_val:,.0f}"  # star winner
row += 2

# ── Winner callout ────────────────────────────────────────────────────────
ws.row_dimensions[row].height = 30
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col-1)
wc = ws.cell(row=row, column=1,
    value="  ★  RECOMMENDATION: Monthly competitions ($150 buy-in × 12/yr) deliver the highest cumulative profit "
          "— 70% more than the Annual model over 4 years. Lower barrier drives higher team volume and commission velocity.")
wc.font = Font(bold=True, size=11, color="065F46")
wc.fill = fill("D1FAE5")
wc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
wc.border = border_thin()


# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — DETAILED P&L per model
# ════════════════════════════════════════════════════════════════════════════
for m in MODELS:
    ws2 = wb.create_sheet(m["name"])
    ws2.sheet_view.showGridLines = False
    colour = MODEL_COLOURS[m["name"]]

    # Title
    ws2.merge_cells("A1:F1")
    tc = ws2["A1"]
    tc.value = f"PUNTING CLUB — {m['name'].upper()} MODEL  |  {m['comps']} comp/yr  ·  ${m['buyin']:,} buy-in  ·  {m['weeks']}wk season"
    tc.font = Font(bold=True, size=14, color=WHITE)
    tc.fill = fill(colour)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32

    # Column headers
    ws2.row_dimensions[3].height = 26
    for ci, txt in enumerate(["", "Year 1", "Year 2", "Year 3", "Year 4"], 1):
        hc = ws2.cell(row=3, column=ci, value=txt)
        hc.font = Font(bold=True, size=11, color=WHITE)
        hc.fill = fill(CARD if ci == 1 else colour)
        hc.alignment = Alignment(horizontal="center", vertical="center")
        hc.border = border_thin()

    ws2.column_dimensions["A"].width = 30
    for ci in range(2, 6): ws2.column_dimensions[get_column_letter(ci)].width = 16

    # Data rows
    ROWS = [
        ("── INPUTS ──────────────────────", None, False, None),
        ("Pubs",              "pubs",          False, "0"),
        ("Competitions/yr",   "comps",         False, "0"),
        ("Buy-in",            "buyin",         False, "$#,##0"),
        ("Teams/pub",         None,            False, "0"),
        ("",                  None,            False, None),
        ("── REVENUE ─────────────────────", None, False, None),
        ("Gross Jackpot",     "jackpot_total",  False, "$#,##0"),
        ("Commission 10%",    "commission",     False, "$#,##0"),
        ("Term Deposit Int.", "interest",       False, "$#,##0"),
        ("Advertising",       "advertising",    False, "$#,##0"),
        ("TOTAL REVENUE",     "total_revenue",  True,  "$#,##0"),
        ("",                  None,             False, None),
        ("── COSTS ───────────────────────", None, False, None),
        ("  Infrastructure",  "infrastructure", False, "$#,##0"),
        ("  HR / Personnel",  "hr",             False, "$#,##0"),
        ("  Marketing",       "marketing",      False, "$#,##0"),
        ("  Legal/Compliance","legal",          False, "$#,##0"),
        ("  Misc",            "misc",           False, "$#,##0"),
        ("  Payment Proc.",   "payment_proc",   False, "$#,##0"),
        ("  Comp. Overhead",  "comp_overhead",  False, "$#,##0"),
        ("TOTAL COSTS",       "total_cost",     True,  "$#,##0"),
        ("",                  None,             False, None),
        ("NET PROFIT",        "net_profit",     True,  "$#,##0"),
        ("Margin %",          "margin",         True,  "0.0%"),
    ]

    r = 4
    for label, key, is_bold, fmt in ROWS:
        ws2.row_dimensions[r].height = 20 if label else 6

        is_section = label.startswith("──")
        is_total = label.startswith("TOTAL") or label == "NET PROFIT" or label == "Margin %"
        row_bg = colour if is_section else \
                 "DCFCE7" if (is_total and label == "NET PROFIT") else \
                 "DBEAFE" if (label == "TOTAL REVENUE") else \
                 "FEE2E2" if (label == "TOTAL COSTS") else \
                 "F8FAFC" if r % 2 == 0 else "FFFFFF"
        fg_col = WHITE if is_section else "111827"

        lc = ws2.cell(row=r, column=1, value=label)
        lc.font = Font(bold=is_bold or is_section, size=10, color=fg_col)
        lc.fill = fill(row_bg)
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border = border_thin()

        if key or fmt:
            for yi, y in enumerate(YEARS):
                val = TEAMS_PER if key is None else results[m["name"]][yi].get(key, 0) if key else ""
                net_neg = (label == "NET PROFIT" and isinstance(val, (int, float)) and val < 0)
                vbg = "FEE2E2" if net_neg else row_bg
                vfg = "991B1B" if net_neg else fg_col
                cell = ws2.cell(row=r, column=2 + yi, value=val)
                cell.number_format = fmt or "General"
                cell.font = Font(bold=is_bold, size=10, color=vfg)
                cell.fill = fill(vbg)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.border = border_thin()
        r += 1

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 — KEY METRICS DASHBOARD
# ════════════════════════════════════════════════════════════════════════════
wd = wb.create_sheet("Key Metrics")
wd.sheet_view.showGridLines = False

wd.merge_cells("A1:P1")
kh = wd["A1"]
kh.value = "PUNTING CLUB — KEY METRICS DASHBOARD"
kh.font = Font(bold=True, size=15, color=WHITE)
kh.fill = fill(DARK_BG)
kh.alignment = Alignment(horizontal="center", vertical="center")
wd.row_dimensions[1].height = 34

# Build a comparison table: revenue, cost, net profit, margin for all models
METRICS = [
    ("Total Revenue", "total_revenue", "$#,##0"),
    ("Total Costs",   "total_cost",    "$#,##0"),
    ("Net Profit",    "net_profit",    "$#,##0"),
    ("Margin %",      "margin",        "0.0%"),
    ("Commission",    "commission",    "$#,##0"),
    ("Interest",      "interest",      "$#,##0"),
]

# Headers: row 3
wd.row_dimensions[3].height = 26
wd.cell(row=3, column=1, value="Metric").font = Font(bold=True, size=11, color=WHITE)
wd["A3"].fill = fill(CARD)
wd["A3"].border = border_thin()
wd["A3"].alignment = Alignment(horizontal="center", vertical="center")
wd.column_dimensions["A"].width = 18

col = 2
for m in MODELS:
    for y in YEARS:
        lbl = f"{m['name'][:3]} Y{y}"
        cell = wd.cell(row=3, column=col, value=lbl)
        cell.font = Font(bold=True, size=10, color=WHITE)
        cell.fill = fill(MODEL_COLOURS[m["name"]])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin()
        wd.column_dimensions[get_column_letter(col)].width = 13
        col += 1

row = 4
for label, key, fmt in METRICS:
    wd.row_dimensions[row].height = 20
    lc = wd.cell(row=row, column=1, value=label)
    lc.font = Font(bold=True, size=10, color="111827")
    lc.fill = fill("F1F5F9")
    lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    lc.border = border_thin()
    col = 2
    for m in MODELS:
        for yi, y in enumerate(YEARS):
            val = results[m["name"]][yi].get(key, 0)
            is_max = all(val >= results[om["name"]][yi].get(key, 0) for om in MODELS) \
                     and key in ("total_revenue","net_profit","commission","margin")
            bg = "D1FAE5" if is_max else ("FEE2E2" if isinstance(val,(int,float)) and val < 0 and key=="net_profit" else "FFFFFF")
            cell = wd.cell(row=row, column=col, value=val)
            cell.number_format = fmt
            cell.font = Font(bold=is_max, size=10, color="065F46" if is_max else "111827")
            cell.fill = fill(bg)
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = border_thin()
            col += 1
    row += 1

# ── Revenue advantage table ───────────────────────────────────────────────
row += 2
wd.row_dimensions[row].height = 24
wd.merge_cells(start_row=row, start_column=1, end_row=row, end_column=17)
tc = wd.cell(row=row, column=1, value="  MONTHLY vs ANNUAL — Revenue Advantage ($)")
tc.font = Font(bold=True, size=12, color=WHITE)
tc.fill = fill(GREEN_D)
tc.alignment = Alignment(horizontal="left", vertical="center")
tc.border = border_thin()
row += 1

wd.row_dimensions[row].height = 20
for ci, hdr in enumerate(["Year 1","Year 2","Year 3","Year 4","Cumulative"], 1):
    cell = wd.cell(row=row, column=ci, value=hdr)
    cell.font = Font(bold=True, size=10, color=WHITE)
    cell.fill = fill(GREEN)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border_thin()
row += 1

wd.row_dimensions[row].height = 22
cumval = 0
for ci, yi in enumerate(range(4), 1):
    diff = results["Monthly"][yi]["net_profit"] - results["Annual"][yi]["net_profit"]
    cumval += diff
    cell = wd.cell(row=row, column=ci, value=diff)
    cell.number_format = "$#,##0"
    cell.font = Font(bold=True, size=11, color="065F46" if diff > 0 else "991B1B")
    cell.fill = fill("D1FAE5" if diff > 0 else "FEE2E2")
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.border = border_thin()

cell = wd.cell(row=row, column=5, value=cumval)
cell.number_format = "$#,##0"
cell.font = Font(bold=True, size=12, color="065F46")
cell.fill = fill("BBF7D0")
cell.alignment = Alignment(horizontal="right", vertical="center")
cell.border = border_thin()


# ════════════════════════════════════════════════════════════════════════════
# SHEET 4 — ASSUMPTIONS & METHODOLOGY
# ════════════════════════════════════════════════════════════════════════════
wa = wb.create_sheet("Assumptions")
wa.sheet_view.showGridLines = False
wa.column_dimensions["A"].width = 35
wa.column_dimensions["B"].width = 55

wa.merge_cells("A1:B1")
ah = wa["A1"]
ah.value = "PUNTING CLUB — ASSUMPTIONS & METHODOLOGY"
ah.font = Font(bold=True, size=14, color=WHITE)
ah.fill = fill(DARK_BG)
ah.alignment = Alignment(horizontal="center", vertical="center")
wa.row_dimensions[1].height = 32

ASSUMPTIONS = [
    ("COMPETITION FORMATS", None),
    ("Annual",              "1 × $1,000 buy-in · 52-week season"),
    ("Bi-Annual",           "2 × $600 buy-in · 26-week seasons"),
    ("Quarterly",           "4 × $350 buy-in · 13-week seasons"),
    ("Monthly",             "12 × $150 buy-in · 4-week seasons"),
    ("", None),
    ("REVENUE ASSUMPTIONS", None),
    ("Commission",          "10% of total gross jackpot collected"),
    ("Term Deposit Rate",   "4% p.a. · held for full competition duration"),
    ("Teams per Pub",       "10 teams average per competition"),
    ("Advertising Revenue", "Y1: $2k · Y2: $10k · Y3: $50k · Y4: $150k (all models same)"),
    ("", None),
    ("COST ASSUMPTIONS", None),
    ("Hosting (Netlify)",   "Y1: $228 · Y2: $1,188 · Y3: $1,188 · Y4: $6,000"),
    ("Database (Supabase)", "Y1: $300 · Y2: $300 · Y3: $7,188 · Y4: $18,000"),
    ("AI / Claude API",     "Y1: $500 · Y2: $1,200 · Y3: $4,000 · Y4: $12,000"),
    ("Domain & SSL",        "Y1–Y3: $50/yr · Y4: $200/yr"),
    ("HR / Personnel",      "Y1: $0 (founder) · Y2: $20k · Y3: $80k · Y4: $240k"),
    ("Marketing",           "Y1: $2k · Y2: $10k · Y3: $30k · Y4: $80k"),
    ("Legal / Compliance",  "Y1: $3k · Y2: $5k · Y3: $10k · Y4: $20k"),
    ("Miscellaneous",       "Y1: $500 · Y2: $2k · Y3: $5k · Y4: $15k"),
    ("Payment Processing",  "1.75% of total jackpot collected (Stripe equivalent)"),
    ("Competition Overhead","$20/comp/pub (Y1) · $15 (Y2) · $10 (Y3) · $8 (Y4) — setup & support"),
    ("", None),
    ("GROWTH ASSUMPTIONS", None),
    ("Pub Growth",          "Y1: 10 pubs · Y2: 30 pubs · Y3: 100 pubs · Y4: 300 pubs"),
    ("Teams per Pub",       "Fixed at 10 — does not model uptick from lower buy-in accessibility"),
    ("Note",                "Monthly model upside is CONSERVATIVE — lower buy-ins likely attract more teams"),
]

r = 3
for label, value in ASSUMPTIONS:
    is_sect = (value is None and label)
    wa.row_dimensions[r].height = 22 if not is_sect else 26

    a_cell = wa.cell(row=r, column=1, value=label)
    b_cell = wa.cell(row=r, column=2, value=value or "")
    a_cell.border = border_thin()
    b_cell.border = border_thin()

    if is_sect:
        a_cell.font = Font(bold=True, size=11, color=WHITE)
        b_cell.font = Font(bold=True, size=11, color=WHITE)
        a_cell.fill = fill(SLATE)
        b_cell.fill = fill(SLATE)
        wa.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    elif not label:
        a_cell.fill = fill("F8FAFC"); b_cell.fill = fill("F8FAFC")
    else:
        bg = "F8FAFC" if r % 2 == 0 else "FFFFFF"
        a_cell.font = Font(bold=True, size=10, color="374151")
        b_cell.font = Font(size=10, color="111827")
        a_cell.fill = fill(bg); b_cell.fill = fill(bg)
        a_cell.alignment = Alignment(horizontal="left", vertical="center")
        b_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    r += 1


# ── Save ──────────────────────────────────────────────────────────────────────
outpath = "/home/user/Punting-club/public/punting-club-financial-model.xlsx"
wb.save(outpath)
print(f"Saved: {outpath}")

# Print summary to verify numbers
print("\n=== NET PROFIT SUMMARY ===")
print(f"{'Model':<12} {'Y1':>12} {'Y2':>12} {'Y3':>12} {'Y4':>12} {'CUM':>14}")
for m in MODELS:
    vals = [results[m["name"]][yi]["net_profit"] for yi in range(4)]
    cum = sum(vals)
    print(f"{m['name']:<12} {vals[0]:>12,.0f} {vals[1]:>12,.0f} {vals[2]:>12,.0f} {vals[3]:>12,.0f} {cum:>14,.0f}")
